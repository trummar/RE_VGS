# importing openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\temp\\roykere_kort_1.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

# Loop will print all values
# of first column
for i in range(1, m_row + 1):
	cell_obj = sheet_obj.cell(row = i, column = 1)
	print(cell_obj.value)

for i in range(1, m_row + 1):
	cell_obj = sheet_obj.cell(row = i, column = 2)
	print(cell_obj.value)


#Lage noe med lister....

	
